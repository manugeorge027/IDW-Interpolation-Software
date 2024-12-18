import pandas as pd
import geopandas as gpd
from shapely.geometry import Point
from openpyxl import load_workbook
from scipy.spatial import distance
import numpy as np
import matplotlib.pyplot as plt
import rasterio
from rasterio.transform import from_origin
from rasterio.mask import mask
import os

# Function to process Excel data
def process_excel(input_file, output_file):
    try:
        # Load Excel file into DataFrame
        df = pd.read_excel(input_file, engine='openpyxl')
        print("Data loaded successfully:")
        print(df)

        # Filter and transpose data for multiple years
        combined_data = pd.DataFrame()
        for val in range(2013, 2022):
            result = df[df.iloc[:, 0] == val]
            trans = result.reset_index(drop=True).T
            trans.columns = [f"S_{val}"]
            combined_data = pd.concat([combined_data, trans], axis=1)

        # Save combined data to Excel without overwriting existing data
        book = load_workbook(output_file)
        sheet = book.active
        for col_idx, column_name in enumerate(combined_data.columns, start=4):
            sheet.cell(row=1, column=col_idx, value=column_name)
            for row_idx, value in enumerate(combined_data[column_name], start=2):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        book.save(output_file)
        print(f"Combined data saved to {output_file}")

    except Exception as e:
        print(f"An error occurred in process_excel: {e}")

# Function to plot Latitude/Longitude to a shapefile
def plot_lat_long_to_shapefile(finput, shapefile, output_shapefile):
    try:
        # Load Excel data
        data = pd.read_excel(finput)
        print("Data from Excel:")
        print(data.head())

        # Validate Latitude and Longitude columns
        if 'Latitude' not in data.columns or 'Longitude' not in data.columns:
            raise ValueError("Latitude and Longitude columns are required in the Excel file.")

        # Convert Latitude and Longitude to numeric
        data['Latitude'] = pd.to_numeric(data['Latitude'], errors='coerce')
        data['Longitude'] = pd.to_numeric(data['Longitude'], errors='coerce')

        # Drop invalid rows
        data = data.dropna(subset=['Latitude', 'Longitude'])

        # Convert to GeoDataFrame
        geometry = [Point(xy) for xy in zip(data['Longitude'], data['Latitude'])]
        gdf = gpd.GeoDataFrame(data, geometry=geometry, crs="EPSG:4326")

        # Load and plot the base shapefile
        base_gdf = gpd.read_file(shapefile)
        ax = base_gdf.plot(color="white", edgecolor="black", figsize=(10, 8))
        gdf.plot(ax=ax, color="red", markersize=5, alpha=0.7, label="Data Points")
        plt.legend()
        plt.title("Latitude and Longitude Points on Shapefile")

        # Save the new GeoDataFrame as a shapefile
        gdf.to_file(output_shapefile, driver="ESRI Shapefile")
        print(f"New shapefile saved to: {output_shapefile}")

    except Exception as e:
        print(f"An error occurred in plot_lat_long_to_shapefile: {e}")

# Function for IDW interpolation
def interpolation_idw(output_shapefile,output_raster_file):
    try:
        # Load shapefile data
        geo_data = gpd.read_file(output_shapefile)

        # Extract coordinates and values
        coords = np.array([(point.x, point.y) for point in geo_data.geometry])
        values = geo_data['S_2013'].astype(float).values  # Ensure 'Value' column exists

        # Generate a grid
        grid_x, grid_y = np.meshgrid(
            np.linspace(coords[:, 0].min(), coords[:, 0].max(), 100),
            np.linspace(coords[:, 1].min(), coords[:, 1].max(), 100)
        )

        # IDW interpolation function
        def idw_interpolation(coords, values, grid_x, grid_y, power=2):
            grid_values = np.zeros_like(grid_x, dtype=float)
            for i in range(grid_x.shape[0]):
                for j in range(grid_x.shape[1]):
                    dist = distance.cdist([(grid_x[i, j], grid_y[i, j])], coords).flatten()
                    if np.any(dist == 0):
                        grid_values[i, j] = values[np.argmin(dist)]
                    else:
                        weights = 1 / (dist ** power)
                        grid_values[i, j] = np.sum(weights * values) / np.sum(weights)
            return grid_values

        # Perform IDW interpolation
        grid_values = idw_interpolation(coords, values, grid_x, grid_y)

        # Plot the results
        plt.figure(figsize=(10, 8))
        plt.contourf(grid_x, grid_y, grid_values, cmap='viridis', levels=20)
        plt.colorbar(label='Interpolated Values')
        plt.scatter(coords[:, 0], coords[:, 1], color='red', label='Data Points', zorder=5)
        plt.title("IDW Interpolation")
        plt.xlabel("Longitude")
        plt.ylabel("Latitude")
        plt.legend()
        plt.show()
        # Ensure the save directory exists on the C drive
      
        

        # Save interpolated grid as a GeoTIFF raster file
        print("Saving IDW interpolation results as a raster file...")
        x_min, x_max = grid_x.min(), grid_x.max()
        y_min, y_max = grid_y.min(), grid_y.max()
        pixel_size_x = (x_max - x_min) / grid_x.shape[1]
        pixel_size_y = (y_max - y_min) / grid_y.shape[0]

        transform = from_origin(x_min, y_max, pixel_size_x, pixel_size_y)
        with rasterio.open(
            output_raster_file,
            'w',
            driver='GTiff',
            height=grid_values.shape[0],
            width=grid_values.shape[1],
            count=1,
            dtype='float32',
            crs="EPSG:4326",
            transform=transform,
        ) as dst:
            dst.write(grid_values, 1)
        print(f"IDW interpolation raster saved to: {output_raster_file}")

    except Exception as e:
        print(f"An error occurred in interpolation_idw: {e}")
def clip_raster_with_shapefile(input_raster, input_shapefile, output_raster):
    
    try:
        # Load the shapefile using GeoPandas
        shapefile = gpd.read_file(input_shapefile)
        print("Shapefile loaded successfully.")
        
        # Ensure the shapefile geometry is in the same CRS as the raster
        with rasterio.open(input_raster) as src:
            raster_crs = src.crs
        if shapefile.crs != raster_crs:
            shapefile = shapefile.to_crs(raster_crs)
            print("Shapefile CRS reprojected to match raster CRS.")

        # Extract the geometry of the shapefile for clipping
        shapes = [feature["geometry"] for feature in shapefile.__geo_interface__["features"]]

        # Open the raster file
        with rasterio.open(input_raster) as src:
            # Mask the raster using the shapefile geometries
            out_image, out_transform = mask(src, shapes, crop=True, nodata=src.nodata)
            out_meta = src.meta.copy()

        # Update metadata for the clipped raster
        out_meta.update({
            "driver": "GTiff",
            "height": out_image.shape[1],
            "width": out_image.shape[2],
            "transform": out_transform
        })

        # Save the clipped raster
        with rasterio.open(output_raster, "w", **out_meta) as dest:
            dest.write(out_image)
            dest.write_mask((out_image[0] != src.nodata).astype("uint8") * 255)  # Ensure correct masking

        print(f"Clipped raster saved successfully to: {output_raster}")

    except Exception as e:
        print(f"An error occurred while clipping the raster: {e}")
# File paths
input_file = 'C:/Users/Hp/Desktop/GIS_Project/projects/data/Seasonal_data.xlsx'
output_file = 'C:/Users/Hp/Desktop/GIS_Project/projects/data/sum.xlsx'
shape_file = 'C:/Users/Hp/Desktop/GIS_Project/projects/shape/KeralaDistrict.shp'
output_shapefile = 'C:/Users/Hp/Desktop/GIS_Project/projects/shape/KeralaDistrict_Points.shp'
output_raster_file='C:/Users/Hp/Desktop/GIS_Project/projects/raster/idw_interpolation.tif'
output_raster='C:/Users/Hp/Desktop/GIS_Project/projects/raster_final/idw_interpolation.tif'

# Execute functions
process_excel(input_file, output_file)
plot_lat_long_to_shapefile(output_file, shape_file, output_shapefile)
interpolation_idw(output_shapefile,output_raster_file)
clip_raster_with_shapefile(output_raster_file, output_shapefile, output_raster)


